# -*- coding: utf-8 -*-
"""
VinaMec Dental Care - User Management UI Automated E2E Test Suite
Tự động hóa kiểm thử giao diện người dùng (UI E2E) sử dụng Playwright và xuất báo cáo Excel.
"""

import os
import sys
import time
import subprocess
import codecs
from datetime import datetime

# Đảm bảo console in UTF-8 không bị lỗi trên Windows
if sys.platform.startswith('win'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())
        sys.stderr = codecs.getwriter("utf-8")(sys.stderr.detach())

# ==============================================================================
# 1. TỰ ĐỘNG KHỞI TẠO VÀ CÀI ĐẶT THƯ VIỆN NẾU CHƯA CÓ
# ==============================================================================
def install_and_import(package, pip_name=None):
    if pip_name is None:
        pip_name = package
    try:
        __import__(package)
    except ImportError:
        print(f"[+] Thư viện '{package}' chưa được cài đặt. Đang tiến hành cài đặt tự động...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
            print(f"[OK] Cài đặt '{package}' thành công!")
        except Exception as e:
            print(f"[LỖI] Không thể cài đặt '{package}': {e}")
            sys.exit(1)

# Cài đặt Playwright & openpyxl nếu chưa có
install_and_import("playwright")
install_and_import("openpyxl")

# Đảm bảo cài đặt trình duyệt của Playwright
print("[+] Đang kiểm tra và cài đặt trình duyệt Chromium của Playwright...")
try:
    subprocess.check_call([sys.executable, "-m", "playwright", "install", "chromium"])
    print("[OK] Trình duyệt Chromium đã sẵn sàng.")
except Exception as e:
    print(f"[CẢNH BÁO] Không thể tự động chạy 'playwright install': {e}")

from playwright.sync_api import sync_playwright, expect
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Cấu hình kiểm thử
BASE_URL = "http://localhost:5173"
ADMIN_EMAIL = "admin@vinamec.vn"
ADMIN_PASSWORD = "admin123"
EXCEL_FILENAME = "user_ui_test_results.xlsx"

# Lưu kết quả thực thi
step_results = {}

# Đặc tả các ca kiểm thử cho Excel
sheets_spec = {
    "UI E2E Cases": [
        {
            "id": "UC_UI_01",
            "description": "Đăng nhập và chuyển hướng trang quản trị",
            "expected_result": "Đăng nhập thành công và tự động chuyển hướng đến Admin Dashboard.",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Mở trang đăng nhập hệ thống.",
                    "step_expected": "Form đăng nhập hiển thị đầy đủ, không gặp lỗi tải trang.",
                    "input_val": f"URL: {BASE_URL}/login"
                },
                {
                    "step_num": 2,
                    "perform": "Điền thông tin và bấm 'Sign In'.",
                    "step_expected": "Đăng nhập thành công và chuyển hướng tới đường dẫn Dashboard /admin.",
                    "input_val": f"Email: {ADMIN_EMAIL}\nPassword: {ADMIN_PASSWORD}"
                }
            ]
        },
        {
            "id": "UC_UI_02",
            "description": "Kiểm tra giao diện trang Quản lý Người dùng",
            "expected_result": "Giao diện tải đầy đủ danh sách và các thẻ thống kê tổng hợp.",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Điều hướng đến trang quản lý người dùng (/admin/users).",
                    "step_expected": "Tiêu đề 'Quản lý Người dùng' hiển thị cùng bảng dữ liệu và thẻ thống kê.",
                    "input_val": "URL: /admin/users"
                }
            ]
        },
        {
            "id": "UC_UI_03",
            "description": "Thêm người dùng mới trên Giao diện",
            "expected_result": "Hội thoại thêm mới mở ra, lưu thành công tài khoản và đóng modal.",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Nhấn nút 'Thêm người dùng', điền Form và click 'Lưu'.",
                    "step_expected": "Modal đóng lại, thông tin được gửi lên máy chủ thành công.",
                    "input_val": "Name: Playwright E2E UI User\nEmail: playwright_e2e_user@example.com\nPhone: 0999999999\nRole: doctor"
                }
            ]
        },
        {
            "id": "UC_UI_04",
            "description": "Tìm kiếm người dùng bằng Thanh tìm kiếm",
            "expected_result": "Bộ lọc hoạt động chính xác, chỉ hiển thị người dùng khớp từ khóa.",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Nhập địa chỉ email vào ô tìm kiếm trên bảng.",
                    "step_expected": "Bảng dữ liệu được lọc động và hàng chứa email vừa tạo hiển thị.",
                    "input_val": "Từ khóa: playwright_e2e_user@example.com"
                }
            ]
        },
        {
            "id": "UC_UI_05",
            "description": "Chỉnh sửa thông tin người dùng trên Giao diện",
            "expected_result": "Cập nhật dữ liệu thành công và hiển thị tức thời trên bảng.",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Nhấn nút 'Sửa', thay đổi Họ tên và click 'Lưu'.",
                    "step_expected": "Dữ liệu được cập nhật thành công lên máy chủ và hiển thị tên mới trên bảng.",
                    "input_val": "Name: Playwright E2E UI User (Updated)"
                }
            ]
        },
        {
            "id": "UC_UI_06",
            "description": "Xóa tài khoản người dùng và xác nhận UI",
            "expected_result": "Xóa thành công, tự động chấp nhận dialog và biến mất khỏi bảng.",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Bấm nút 'Xóa', đồng ý với hộp thoại xác nhận của trình duyệt.",
                    "step_expected": "Trình duyệt xóa tài khoản, bảng danh sách tải lại và không còn hiển thị tài khoản.",
                    "input_val": "Đồng ý popup confirm"
                }
            ]
        }
    ]
}

def record_result(tc_id, step_num, success, message):
    sheet_name = "UI E2E Cases"
    step_results[(sheet_name, tc_id, step_num)] = (success, message)

def run_ui_test():
    print("======================================================================")
    print("BẮT ĐẦU CHẠY PLAYWRIGHT UI TEST CHO QUẢN LÝ NGƯỜI DÙNG (USER MANAGEMENT)")
    print("======================================================================")
    
    # Cấu hình chụp ảnh
    os.makedirs("test/screenshots", exist_ok=True)
    
    with sync_playwright() as p:
        print("[*] Đang khởi chạy trình duyệt Chromium...")
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        try:
            # --- UC_UI_01: Đăng nhập ---
            print("\n👉 [UC_UI_01 - Bước 1] Truy cập trang đăng nhập...")
            try:
                page.goto(f"{BASE_URL}/login")
                page.wait_for_selector("input[type='email']", timeout=5000)
                page.screenshot(path="test/screenshots/1_login_page.png")
                record_result("UC_UI_01", 1, True, "Truy cập trang login thành công, các trường nhập liệu hiển thị.")
            except Exception as e:
                record_result("UC_UI_01", 1, False, f"Không thể tải trang đăng nhập: {e}. Vui lòng chạy frontend.")
                raise e

            print("\n👉 [UC_UI_01 - Bước 2] Điền thông tin đăng nhập...")
            try:
                page.fill("input[type='email']", ADMIN_EMAIL)
                page.fill("input[type='password']", ADMIN_PASSWORD)
                page.screenshot(path="test/screenshots/2_login_filled.png")
                page.click("button:has-text('Sign In')")
                # Đợi redirect sang admin dashboard
                page.wait_for_url("**/admin", timeout=5000)
                record_result("UC_UI_01", 2, True, "Đăng nhập thành công, chuyển hướng đến /admin.")
            except Exception as e:
                record_result("UC_UI_01", 2, False, f"Đăng nhập thất bại: {e}. Vui lòng khởi động backend.")
                raise e

            # --- UC_UI_02: Điều hướng Quản lý Người dùng ---
            print("\n👉 [UC_UI_02 - Bước 1] Điều hướng đến trang quản lý người dùng...")
            try:
                page.goto(f"{BASE_URL}/admin/users")
                page.wait_for_selector("h1:has-text('Quản lý Người dùng')", timeout=5000)
                page.screenshot(path="test/screenshots/3_admin_users.png")
                record_result("UC_UI_02", 1, True, "Tải trang quản lý người dùng thành công, hiển thị đầy đủ UI.")
            except Exception as e:
                record_result("UC_UI_02", 1, False, f"Tải trang quản lý người dùng thất bại: {e}")
                raise e

            # --- UC_UI_03: Thêm người dùng mới ---
            print("\n👉 [UC_UI_03 - Bước 1] Điền form thêm người dùng mới...")
            test_name = "Playwright E2E UI User"
            test_email = "playwright_e2e_user@example.com"
            test_phone = "0999999999"
            try:
                page.click("button:has-text('Thêm người dùng')")
                page.wait_for_selector("input[placeholder='Nhập họ và tên']", timeout=3000)
                
                page.fill("input[placeholder='Nhập họ và tên']", test_name)
                page.fill("input[placeholder='Nhập địa chỉ email']", test_email)
                page.fill("input[placeholder='Nhập số điện thoại']", test_phone)
                page.select_option("select", value="doctor")
                
                page.screenshot(path="test/screenshots/4_add_user_form.png")
                page.click("button[type='submit']:has-text('Lưu')")
                page.wait_for_selector("input[placeholder='Nhập họ và tên']", state="detached", timeout=5000)
                record_result("UC_UI_03", 1, True, f"Tạo người dùng mới '{test_name}' thành công.")
            except Exception as e:
                record_result("UC_UI_03", 1, False, f"Không tạo được người dùng: {e}")
                raise e

            # --- UC_UI_04: Tìm kiếm người dùng ---
            print("\n👉 [UC_UI_04 - Bước 1] Tìm kiếm người dùng vừa tạo...")
            try:
                search_input = page.locator("input[placeholder='Tìm kiếm người dùng...']")
                search_input.fill(test_email)
                time.sleep(1)
                page.screenshot(path="test/screenshots/5_search_result.png")
                
                user_row = page.locator(f"tr:has-text('{test_email}')")
                expect(user_row).to_be_visible(timeout=3000)
                record_result("UC_UI_04", 1, True, "Thanh tìm kiếm hoạt động tốt, hiển thị kết quả lọc chính xác.")
            except Exception as e:
                record_result("UC_UI_04", 1, False, f"Lọc tìm kiếm không hiển thị người dùng: {e}")
                raise e

            # --- UC_UI_05: Chỉnh sửa thông tin ---
            print("\n👉 [UC_UI_05 - Bước 1] Sửa thông tin họ tên người dùng...")
            updated_name = "Playwright E2E UI User (Updated)"
            try:
                user_row.locator("button:has-text('Sửa')").click()
                page.wait_for_selector("input[placeholder='Nhập họ và tên']", timeout=3000)
                page.fill("input[placeholder='Nhập họ và tên']", updated_name)
                page.screenshot(path="test/screenshots/6_edit_user_form.png")
                page.click("button[type='submit']:has-text('Lưu')")
                page.wait_for_selector("input[placeholder='Nhập họ và tên']", state="detached", timeout=5000)
                
                # Tìm lại xem đã sửa chưa
                search_input.fill("")
                search_input.fill(updated_name)
                time.sleep(1)
                
                updated_row = page.locator(f"tr:has-text('{updated_name}')")
                expect(updated_row).to_be_visible(timeout=3000)
                record_result("UC_UI_05", 1, True, f"Cập nhật thành công họ tên thành: '{updated_name}'.")
            except Exception as e:
                record_result("UC_UI_05", 1, False, f"Chỉnh sửa thông tin thất bại: {e}")
                raise e

            # --- UC_UI_06: Xóa người dùng ---
            print("\n👉 [UC_UI_06 - Bước 1] Thực thi xóa và confirm dialog...")
            try:
                # Tự động đồng ý khi hiện confirm dialog
                page.on("dialog", lambda dialog: dialog.accept())
                
                updated_row.locator("button:has-text('Xóa')").click()
                time.sleep(1)
                
                # Check xem đã mất chưa
                search_input.fill("")
                search_input.fill(updated_name)
                time.sleep(1)
                expect(page.locator(f"tr:has-text('{updated_name}')")).to_have_count(0, timeout=3000)
                page.screenshot(path="test/screenshots/7_after_deletion.png")
                record_result("UC_UI_06", 1, True, "Đã xóa người dùng và cập nhật UI biến mất khỏi bảng thành công.")
            except Exception as e:
                record_result("UC_UI_06", 1, False, f"Xóa người dùng thất bại: {e}")
                raise e

        except Exception as overall_error:
            print(f"\n❌ [LỖI TOÀN CỤC]: {overall_error}")
            page.screenshot(path="test/screenshots/error_exception.png")
            # Đánh dấu những bước chưa thực thi hoặc lỗi
            for tc in sheets_spec["UI E2E Cases"]:
                for step in tc["steps"]:
                    key = ("UI E2E Cases", tc["id"], step["step_num"])
                    if key not in step_results:
                        step_results[key] = (False, f"Bỏ qua hoặc lỗi do tiến trình kiểm thử đứt gãy: {overall_error}")
        finally:
            context.close()
            browser.close()

            # Tính toán trạng thái Pass/Fail cho từng TC
            for tc in sheets_spec["UI E2E Cases"]:
                tc_id = tc["id"]
                tc_passed = True
                for step in tc["steps"]:
                    key = ("UI E2E Cases", tc_id, step["step_num"])
                    s_res = step_results.get(key)
                    if not s_res or not s_res[0]:
                        tc_passed = False
                        break
                tc["status"] = "Pass" if tc_passed else "Fail"

            # Xuất kết quả ra Excel
            export_excel()

# ==============================================================================
# 4. HÀM XUẤT EXCEL STYLED CAO CẤP
# ==============================================================================
def export_excel():
    print(f"\n[+] Đang tạo báo cáo Excel kết quả kiểm thử UI tại: {EXCEL_FILENAME}...")
    wb = Workbook()
    
    COLOR_HEADER_BG = "E2EFDA"     # Màu xanh lá cây nhạt (Light Sage Green)
    COLOR_BORDER = "CCCCCC"        # Viền xám nhạt
    COLOR_PASS_FILL = "00B050"     # Màu xanh lục tươi
    COLOR_FAIL_FILL = "FF0000"     # Màu đỏ tươi
    COLOR_WHITE = "FFFFFF"
    
    font_header_main = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_header_sub = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_body = Font(name="Segoe UI", size=9.5, color="000000")
    font_bold_body = Font(name="Segoe UI", size=9.5, bold=True, color="000000")
    
    font_status_pass = Font(name="Segoe UI", size=10, bold=True, color=COLOR_WHITE)
    font_status_fail = Font(name="Segoe UI", size=10, bold=True, color=COLOR_WHITE)
    font_yellow_banner = Font(name="Segoe UI", size=9.5, bold=True, color="000000")
    
    fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    fill_pass = PatternFill(start_color=COLOR_PASS_FILL, end_color=COLOR_PASS_FILL, fill_type="solid")
    fill_fail = PatternFill(start_color=COLOR_FAIL_FILL, end_color=COLOR_FAIL_FILL, fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    thin_side = Side(border_style="thin", color=COLOR_BORDER)
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center_both = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_both = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    is_first = True
    for sheet_name, tc_list in sheets_spec.items():
        if is_first:
            ws = wb.active
            ws.title = sheet_name
            is_first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
            
        ws.views.sheetView[0].showGridLines = True
        
        # Banner thời gian
        ws.merge_cells("G1:H1")
        banner_cell = ws["G1"]
        banner_cell.value = f"Ngày kiểm thử UI : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        banner_cell.font = font_yellow_banner
        banner_cell.fill = fill_yellow
        banner_cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=1, column=7).border = border_cell
        ws.cell(row=1, column=8).border = border_cell
        
        # Tiêu đề bảng
        ws.merge_cells("A3:A4") # Test Case ID
        ws.merge_cells("B3:B4") # Test Case Description
        ws.merge_cells("C3:D3") # Test Procedures
        ws.merge_cells("E3:E4") # Test Case Input Value
        ws.merge_cells("F3:F4") # Test Case Expected Result
        ws.merge_cells("G3:G4") # Status
        
        ws["A3"] = "Test Case ID"
        ws["B3"] = "Test Case Description"
        ws["C3"] = "Test Procedures"
        ws["E3"] = "Test Case Input Value"
        ws["F3"] = "Test Case Expected Result"
        ws["G3"] = "Status"
        
        ws["C4"] = "Step to Perform"
        ws["D4"] = "Step Expected Result / Actual Result"
        
        for col in range(1, 8):
            cell_r3 = ws.cell(row=3, column=col)
            cell_r4 = ws.cell(row=4, column=col)
            
            cell_r3.font = font_header_main
            cell_r3.fill = fill_header
            cell_r3.border = border_cell
            cell_r3.alignment = align_center_both
            
            cell_r4.font = font_header_sub
            cell_r4.fill = fill_header
            cell_r4.border = border_cell
            cell_r4.alignment = align_center_both

        ws.row_dimensions[3].height = 24
        ws.row_dimensions[4].height = 24

        # Đổ dữ liệu
        current_row = 5
        for tc in tc_list:
            steps_count = len(tc["steps"])
            start_merge_row = current_row
            end_merge_row = current_row + steps_count - 1
            
            ws.cell(row=current_row, column=1, value=tc["id"])
            ws.cell(row=current_row, column=2, value=tc["description"])
            ws.cell(row=current_row, column=6, value=tc["expected_result"])
            ws.cell(row=current_row, column=7, value=tc["status"])
            
            for idx, step in enumerate(tc["steps"]):
                row_idx = current_row + idx
                
                step_num_perform = f"{step['step_num']}. {step['perform']}"
                ws.cell(row=row_idx, column=3, value=step_num_perform)
                
                key = (sheet_name, tc["id"], step["step_num"])
                s_res = step_results.get(key)
                actual_text = s_res[1] if s_res else "Chưa thực thi"
                
                combined_expected_actual = f"Mong đợi: {step['step_expected']}\nThực tế: {actual_text}"
                ws.cell(row=row_idx, column=4, value=combined_expected_actual)
                ws.cell(row=row_idx, column=5, value=step["input_val"])
                
                for col in range(1, 8):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = border_cell
                    cell.font = font_body
                    if col in [1, 7]:
                        cell.alignment = align_center_both
                    else:
                        cell.alignment = align_left_both
                        
                ws.row_dimensions[row_idx].height = 65
                
            ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=end_merge_row, end_column=1)
            ws.merge_cells(start_row=start_merge_row, start_column=2, end_row=end_merge_row, end_column=2)
            ws.merge_cells(start_row=start_merge_row, start_column=6, end_row=end_merge_row, end_column=6)
            ws.merge_cells(start_row=start_merge_row, start_column=7, end_row=end_merge_row, end_column=7)
            
            ws.cell(row=start_merge_row, column=1).font = font_bold_body
            ws.cell(row=start_merge_row, column=1).alignment = align_center_both
            ws.cell(row=start_merge_row, column=2).font = font_bold_body
            ws.cell(row=start_merge_row, column=6).alignment = align_left_both
            
            status_val = tc["status"]
            for r in range(start_merge_row, end_merge_row + 1):
                cell_status = ws.cell(row=r, column=7)
                if status_val == "Pass":
                    cell_status.fill = fill_pass
                    cell_status.font = font_status_pass
                else:
                    cell_status.fill = fill_fail
                    cell_status.font = font_status_fail
                cell_status.alignment = align_center_both
                cell_status.border = border_cell
                
            current_row += steps_count

        column_widths = {1: 15, 2: 25, 3: 35, 4: 55, 5: 25, 6: 30, 7: 15}
        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

    try:
        wb.save(EXCEL_FILENAME)
        print(f"[OK] Đã xuất báo cáo Excel thành công: '{os.path.abspath(EXCEL_FILENAME)}'")
    except Exception as e:
        print(f"[ERROR] Không thể lưu file Excel: {e}")

if __name__ == "__main__":
    run_ui_test()
