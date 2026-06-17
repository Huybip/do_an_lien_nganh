# -*- coding: utf-8 -*-
"""
VinaMec Dental Care - User Management API Use Case Test Suite
Tự động hóa kiểm thử các kịch bản nghiệp vụ (Use Cases) của hệ thống Quản lý người dùng và xuất báo cáo Excel.
"""

import os
import sys
import json
import subprocess
import codecs
from datetime import datetime

# Đảm bảo console in UTF-8 không bị lỗi trên Windows
if sys.platform.startswith('win'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())
        sys.stderr = codecs.getwriter("utf-8")(sys.stderr.detach())

# ==============================================================================
# 1. TỰ ĐỘNG KHỞI TẠO VÀ CÀI ĐẶT THƯ VIỆN NẾU CHƯA CÓ
# ==============================================================================
def install_and_import(package, pip_name=None):
    if pip_name is None:
        pip_name = package
    try:
        __import__(package)
    except ImportError:
        print(f"[+] Thư viện '{package}' chưa được cài đặt. Đang tiến hành cài đặt tự động...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
            print(f"[OK] Cài đặt '{package}' thành công!")
        except Exception as e:
            print(f"[LỖI] Không thể cài đặt '{package}': {e}")
            sys.exit(1)

install_and_import("requests")
install_and_import("openpyxl")

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# 2. CẤU HÌNH LIÊN KẾT HỆ THỐNG
# ==============================================================================
BASE_URL = "http://localhost:5000/api"
ADMIN_CREDENTIALS = {"email": "admin@vinamec.vn", "password": "admin123"}
EXCEL_FILENAME = "user_uc_test_results.xlsx"

# State lưu trữ giữa các bước kiểm thử
state = {
    "token": None,
    "created_user_id": None,
    "created_user_email": "test_uc_doctor@vinamec.vn",
    "doctor_profile_id": None
}

# Lưu kết quả thực thi chi tiết của từng Step để xuất Excel:
# key: (sheet_name, tc_id, step_num) -> (success_boolean, actual_message)
step_results = {}

# Định nghĩa Đặc tả các Test Cases để tạo cấu trúc Excel chuẩn
sheets_spec = {
    "API Use Cases": [
        {
            "id": "UC_API_01",
            "description": "Đăng nhập tài khoản Quản trị viên (Admin Login)",
            "expected_result": "Đăng nhập thành công, hệ thống trả về Token quyền Admin (HTTP 200).",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Gửi yêu cầu POST đăng nhập đến /auth/login.",
                    "step_expected": "Đăng nhập thành công, trả về JWT token để gọi các API Admin khác.",
                    "input_val": f"Email: {ADMIN_CREDENTIALS['email']}\nPassword: {ADMIN_CREDENTIALS['password']}"
                }
            ]
        },
        {
            "id": "UC_API_02",
            "description": "Truy xuất danh sách người dùng trong hệ thống",
            "expected_result": "Trả về danh sách tất cả các tài khoản người dùng thành công (HTTP 200).",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Gửi yêu cầu GET đến /users kèm Admin Token.",
                    "step_expected": "Trả về mã HTTP 200 và danh sách người dùng trong mảng data.",
                    "input_val": "Authorization Bearer Token"
                }
            ]
        },
        {
            "id": "UC_API_03",
            "description": "Tạo mới tài khoản người dùng với vai trò Bác sĩ",
            "expected_result": "Tạo tài khoản thành công (HTTP 201). Hệ thống tự động tạo Doctor Profile.",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Gửi yêu cầu POST tạo user đến /users với vai trò doctor.",
                    "step_expected": "Đăng ký thành công, trả về HTTP 201 và thông tin User mới tạo.",
                    "input_val": f"Name: Bác Sĩ Nghiệp Vụ Test\nEmail: {state['created_user_email']}\nRole: doctor\nSpecialization: Orthodontics"
                },
                {
                    "step_num": 2,
                    "perform": "Gửi yêu cầu GET đến /doctors để kiểm tra Profile tự động tạo.",
                    "step_expected": "Tìm thấy Doctor Profile liên kết chính xác với User ID vừa được tạo mới.",
                    "input_val": "Kiểm tra User ID khớp trong danh sách doctors"
                }
            ]
        },
        {
            "id": "UC_API_04",
            "description": "Ngăn chặn tạo tài khoản trùng Email",
            "expected_result": "Hệ thống từ chối tạo và trả về mã lỗi HTTP 409 Conflict.",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Gửi yêu cầu POST tạo user trùng lặp email vừa tạo.",
                    "step_expected": "Trả về HTTP 409 và thông báo lỗi Email already exists.",
                    "input_val": f"Email: {state['created_user_email']}"
                }
            ]
        },
        {
            "id": "UC_API_05",
            "description": "Cập nhật thông tin tài khoản người dùng",
            "expected_result": "Cập nhật thông tin thành công và lưu vào DB (HTTP 200).",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Gửi yêu cầu PUT đến /users/:id để thay đổi Tên và Số điện thoại.",
                    "step_expected": "Trả về HTTP 200 và thông tin User đã sửa đổi thành công.",
                    "input_val": "Name: Bác Sĩ Nghiệp Vụ Test (Đã cập nhật)\nPhone: 0988888888"
                }
            ]
        },
        {
            "id": "UC_API_06",
            "description": "Bật/Tắt kích hoạt trạng thái hoạt động tài khoản",
            "expected_result": "Thay đổi trạng thái isActive thành công (HTTP 200).",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Gửi yêu cầu PUT đến /users/:id/toggle-active.",
                    "step_expected": "Trả về HTTP 200 và giá trị isActive được đảo ngược thành công.",
                    "input_val": "User ID của tài khoản vừa tạo"
                }
            ]
        },
        {
            "id": "UC_API_07",
            "description": "Ngăn chặn Admin tự xóa tài khoản của chính mình",
            "expected_result": "Hệ thống từ chối xóa và trả về mã lỗi HTTP 400 Bad Request.",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Admin gửi yêu cầu DELETE đến /users/:my_admin_id.",
                    "step_expected": "Trả về HTTP 400 và thông báo không thể tự xóa tài khoản chính mình.",
                    "input_val": f"Admin Email: {ADMIN_CREDENTIALS['email']}"
                }
            ]
        },
        {
            "id": "UC_API_08",
            "description": "Xóa tài khoản người dùng khỏi hệ thống",
            "expected_result": "Xóa thành công tài khoản người dùng thử nghiệm (HTTP 200).",
            "status": "Pass",
            "steps": [
                {
                    "step_num": 1,
                    "perform": "Gửi yêu cầu DELETE đến /users/:id.",
                    "step_expected": "Trả về HTTP 200 và xác nhận đã xóa thành công.",
                    "input_val": "User ID của tài khoản vừa tạo"
                },
                {
                    "step_num": 2,
                    "perform": "Gửi yêu cầu GET đến /users/:id để xác thực trạng thái.",
                    "step_expected": "Trả về HTTP 404 và báo lỗi User not found.",
                    "input_val": "User ID đã bị xóa"
                }
            ]
        }
    ]
}

def record_result(tc_id, step_num, success, message):
    sheet_name = "API Use Cases"
    step_results[(sheet_name, tc_id, step_num)] = (success, message)

def run_api_step(tc_id, step_num, method, endpoint, payload=None, headers=None, expected_status=200):
    url = f"{BASE_URL}{endpoint}"
    print(f"\n⚡ [{tc_id} - Bước {step_num}] {method.upper()} {url}")
    if payload:
        p_str = json.dumps(payload, ensure_ascii=False)
        print(f"   👉 Body: {p_str[:120]}..." if len(p_str) > 120 else f"   👉 Body: {p_str}")
    
    try:
        if method.upper() == "GET":
            response = requests.get(url, headers=headers, timeout=10)
        elif method.upper() == "POST":
            response = requests.post(url, json=payload, headers=headers, timeout=10)
        elif method.upper() == "PUT":
            response = requests.put(url, json=payload, headers=headers, timeout=10)
        elif method.upper() == "DELETE":
            response = requests.delete(url, headers=headers, timeout=10)
        else:
            msg = f"Unsupported HTTP method: {method}"
            record_result(tc_id, step_num, False, msg)
            return False, msg

        status_code = response.status_code
        try:
            res_data = response.json()
        except:
            res_data = None

        if status_code == expected_status:
            msg = f"HTTP {status_code}: Thành công."
            if res_data and res_data.get("message"):
                msg += f" - {res_data.get('message')}"
            print(f" ✅ [OK] {msg}")
            record_result(tc_id, step_num, True, msg)
            return True, res_data
        else:
            err_msg = res_data.get("message") if res_data else response.text
            msg = f"HTTP {status_code} (Mong đợi {expected_status}): {err_msg}"
            print(f" ❌ [FAIL] {msg}")
            record_result(tc_id, step_num, False, msg)
            return False, res_data
    except Exception as e:
        msg = f"Exception: {e}"
        print(f" ❌ [EXCEPTION] {msg}")
        record_result(tc_id, step_num, False, msg)
        return False, str(e)

# ==============================================================================
# 3. KỊCH BẢN CHẠY KIỂM THỬ (USE CASES)
# ==============================================================================
def main():
    print("======================================================================")
    print("BẮT ĐẦU CHẠY API USE CASE TEST CHO QUẢN LÝ NGƯỜI DÙNG (USER MANAGEMENT)")
    print("======================================================================")
    
    # 0. Dọn dẹp trước nếu còn dữ liệu rác (tự phục hồi - Self-Healing)
    print("[*] Chuẩn bị môi trường: Xóa người dùng thử nghiệm cũ (nếu có)")
    try:
        r = requests.post(f"{BASE_URL}/auth/login", json=ADMIN_CREDENTIALS, timeout=5)
        if r.status_code == 200:
            tok = r.json()["data"]["token"]
            hd = {"Authorization": f"Bearer {tok}"}
            rl = requests.get(f"{BASE_URL}/users?search={state['created_user_email']}", headers=hd, timeout=5)
            if rl.status_code == 200 and rl.json().get("data"):
                for u in rl.json()["data"]:
                    requests.delete(f"{BASE_URL}/users/{u['_id']}", headers=hd, timeout=5)
                    print(f"    [-] Đã dọn dẹp tài khoản cũ: {u['email']}")
    except Exception as e:
        print(f"    [!] Bỏ qua dọn dẹp lỗi: {e}")

    # --- UC_API_01: Đăng nhập quyền Admin ---
    success, res = run_api_step("UC_API_01", 1, "POST", "/auth/login", payload=ADMIN_CREDENTIALS, expected_status=200)
    if not success or not res or not res.get("success"):
        print("\n❌ Thất bại ngay bước đăng nhập Admin. Dừng chạy test và xuất báo cáo Excel lỗi.")
        # Đánh dấu lỗi cho toàn bộ các case còn lại
        for tc in sheets_spec["API Use Cases"]:
            tc["status"] = "Fail"
            for step in tc["steps"]:
                key = ("API Use Cases", tc["id"], step["step_num"])
                if key not in step_results:
                    step_results[key] = (False, "Chưa thực thi do bước đăng nhập Admin thất bại. Vui lòng bật backend.")
        export_excel()
        sys.exit(1)
        
    state["token"] = res["data"]["token"]
    headers = {"Authorization": f"Bearer {state['token']}"}

    # --- UC_API_02: Lấy danh sách người dùng ---
    success, res = run_api_step("UC_API_02", 1, "GET", "/users", headers=headers, expected_status=200)

    # --- UC_API_03: Tạo mới người dùng (Bác sĩ) ---
    new_user = {
        "name": "Bác Sĩ Nghiệp Vụ Test",
        "email": state["created_user_email"],
        "password": "DoctorPassword@123",
        "role": "doctor",
        "phone": "0912345678",
        "specialization": "Orthodontics"
    }
    success, res = run_api_step("UC_API_03", 1, "POST", "/users", payload=new_user, headers=headers, expected_status=201)
    if success and res and res.get("data"):
        state["created_user_id"] = res["data"]["_id"]
        
        # Bước 2: Kiểm tra Doctor Profile tự tạo
        try:
            doc_res = requests.get(f"{BASE_URL}/doctors", headers=headers, timeout=10)
            if doc_res.status_code == 200:
                doctors = doc_res.json().get("data", [])
                found = False
                for d in doctors:
                    u_id = d["user"]["_id"] if isinstance(d["user"], dict) else d["user"]
                    if u_id == state["created_user_id"]:
                        state["doctor_profile_id"] = d["_id"]
                        found = True
                        break
                if found:
                    msg = f"Đã tự động tạo Doctor Profile. ID Profile: {state['doctor_profile_id']}"
                    record_result("UC_API_03", 2, True, msg)
                else:
                    msg = "Không tìm thấy Doctor Profile liên kết với user mới tạo."
                    record_result("UC_API_03", 2, False, msg)
            else:
                msg = f"Không lấy được danh sách bác sĩ để đối soát. HTTP {doc_res.status_code}"
                record_result("UC_API_03", 2, False, msg)
        except Exception as e:
            record_result("UC_API_03", 2, False, f"Lỗi ngoại lệ: {e}")
    else:
        # Nếu bước 1 lỗi, bước 2 coi như Fail
        record_result("UC_API_03", 2, False, "Bỏ qua do bước tạo user lỗi.")

    # --- UC_API_04: Kiểm tra tạo trùng email ---
    run_api_step("UC_API_04", 1, "POST", "/users", payload=new_user, headers=headers, expected_status=409)

    # --- UC_API_05: Cập nhật thông tin người dùng ---
    if state["created_user_id"]:
        update_payload = {
            "name": "Bác Sĩ Nghiệp Vụ Test (Đã cập nhật)",
            "phone": "0988888888"
        }
        run_api_step("UC_API_05", 1, "PUT", f"/users/{state['created_user_id']}", payload=update_payload, headers=headers, expected_status=200)
    else:
        record_result("UC_API_05", 1, False, "Bỏ qua do không có User ID thử nghiệm.")

    # --- UC_API_06: Bật/Tắt kích hoạt tài khoản ---
    if state["created_user_id"]:
        run_api_step("UC_API_06", 1, "PUT", f"/users/{state['created_user_id']}/toggle-active", headers=headers, expected_status=200)
    else:
        record_result("UC_API_06", 1, False, "Bỏ qua do không có User ID thử nghiệm.")

    # --- UC_API_07: Admin tự xóa tài khoản của chính mình ---
    admin_user_id = None
    try:
        me_res = requests.get(f"{BASE_URL}/users", headers=headers, timeout=5)
        if me_res.status_code == 200:
            users_list = me_res.json().get("data", [])
            for u in users_list:
                if u["email"] == ADMIN_CREDENTIALS["email"]:
                    admin_user_id = u["_id"]
                    break
    except Exception as e:
        print(f" ⚠️ Không lấy được Admin ID để test tự xóa: {e}")

    if admin_user_id:
        run_api_step("UC_API_07", 1, "DELETE", f"/users/{admin_user_id}", headers=headers, expected_status=400)
    else:
        record_result("UC_API_07", 1, False, "Bỏ qua vì không lấy được Admin ID để test.")

    # --- UC_API_08: Xóa người dùng thử nghiệm thành công & Check 404 ---
    if state["created_user_id"]:
        run_api_step("UC_API_08", 1, "DELETE", f"/users/{state['created_user_id']}", headers=headers, expected_status=200)
        run_api_step("UC_API_08", 2, "GET", f"/users/{state['created_user_id']}", headers=headers, expected_status=404)
    else:
        record_result("UC_API_08", 1, False, "Bỏ qua do không tạo được User trước đó.")
        record_result("UC_API_08", 2, False, "Bỏ qua do không tạo được User trước đó.")

    # Cập nhật trạng thái tổng thể cho các Test Cases trong Excel sheet_spec
    for tc in sheets_spec["API Use Cases"]:
        tc_id = tc["id"]
        tc_passed = True
        for step in tc["steps"]:
            key = ("API Use Cases", tc_id, step["step_num"])
            s_res = step_results.get(key)
            if not s_res or not s_res[0]:
                tc_passed = False
                break
        tc["status"] = "Pass" if tc_passed else "Fail"

    # Xuất báo cáo Excel
    export_excel()

# ==============================================================================
# 4. HÀM XUẤT EXCEL STYLED CAO CẤP (REUSED TỪ TEST_API)
# ==============================================================================
def export_excel():
    print(f"\n[+] Đang tạo báo cáo Excel kết quả kiểm thử tại: {EXCEL_FILENAME}...")
    wb = Workbook()
    
    # Palette Màu Sắc Giao Diện Premium
    COLOR_HEADER_BG = "E2EFDA"     # Màu xanh lá cây nhạt (Light Sage Green)
    COLOR_BORDER = "CCCCCC"        # Viền xám nhạt
    COLOR_PASS_FILL = "00B050"     # Màu xanh lục tươi
    COLOR_FAIL_FILL = "FF0000"     # Màu đỏ tươi
    COLOR_WHITE = "FFFFFF"
    
    font_header_main = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_header_sub = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_body = Font(name="Segoe UI", size=9.5, color="000000")
    font_bold_body = Font(name="Segoe UI", size=9.5, bold=True, color="000000")
    
    font_status_pass = Font(name="Segoe UI", size=10, bold=True, color=COLOR_WHITE)
    font_status_fail = Font(name="Segoe UI", size=10, bold=True, color=COLOR_WHITE)
    font_yellow_banner = Font(name="Segoe UI", size=9.5, bold=True, color="000000")
    
    fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    fill_pass = PatternFill(start_color=COLOR_PASS_FILL, end_color=COLOR_PASS_FILL, fill_type="solid")
    fill_fail = PatternFill(start_color=COLOR_FAIL_FILL, end_color=COLOR_FAIL_FILL, fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    thin_side = Side(border_style="thin", color=COLOR_BORDER)
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center_both = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_both = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    is_first = True
    for sheet_name, tc_list in sheets_spec.items():
        if is_first:
            ws = wb.active
            ws.title = sheet_name
            is_first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
            
        ws.views.sheetView[0].showGridLines = True
        
        # Banner thời gian
        ws.merge_cells("G1:H1")
        banner_cell = ws["G1"]
        banner_cell.value = f"Ngày kiểm thử : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        banner_cell.font = font_yellow_banner
        banner_cell.fill = fill_yellow
        banner_cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=1, column=7).border = border_cell
        ws.cell(row=1, column=8).border = border_cell
        
        # Tiêu đề bảng
        ws.merge_cells("A3:A4") # Test Case ID
        ws.merge_cells("B3:B4") # Test Case Description
        ws.merge_cells("C3:D3") # Test Procedures
        ws.merge_cells("E3:E4") # Test Case Input Value
        ws.merge_cells("F3:F4") # Test Case Expected Result
        ws.merge_cells("G3:G4") # Status
        
        ws["A3"] = "Test Case ID"
        ws["B3"] = "Test Case Description"
        ws["C3"] = "Test Procedures"
        ws["E3"] = "Test Case Input Value"
        ws["F3"] = "Test Case Expected Result"
        ws["G3"] = "Status"
        
        ws["C4"] = "Step to Perform"
        ws["D4"] = "Step Expected Result / Actual Result"
        
        for col in range(1, 8):
            cell_r3 = ws.cell(row=3, column=col)
            cell_r4 = ws.cell(row=4, column=col)
            
            cell_r3.font = font_header_main
            cell_r3.fill = fill_header
            cell_r3.border = border_cell
            cell_r3.alignment = align_center_both
            
            cell_r4.font = font_header_sub
            cell_r4.fill = fill_header
            cell_r4.border = border_cell
            cell_r4.alignment = align_center_both

        ws.row_dimensions[3].height = 24
        ws.row_dimensions[4].height = 24

        # Đổ dữ liệu
        current_row = 5
        for tc in tc_list:
            steps_count = len(tc["steps"])
            start_merge_row = current_row
            end_merge_row = current_row + steps_count - 1
            
            ws.cell(row=current_row, column=1, value=tc["id"])
            ws.cell(row=current_row, column=2, value=tc["description"])
            ws.cell(row=current_row, column=6, value=tc["expected_result"])
            ws.cell(row=current_row, column=7, value=tc["status"])
            
            for idx, step in enumerate(tc["steps"]):
                row_idx = current_row + idx
                
                step_num_perform = f"{step['step_num']}. {step['perform']}"
                ws.cell(row=row_idx, column=3, value=step_num_perform)
                
                key = (sheet_name, tc["id"], step["step_num"])
                s_res = step_results.get(key)
                actual_text = s_res[1] if s_res else "Chưa thực thi"
                
                combined_expected_actual = f"Mong đợi: {step['step_expected']}\nThực tế: {actual_text}"
                ws.cell(row=row_idx, column=4, value=combined_expected_actual)
                ws.cell(row=row_idx, column=5, value=step["input_val"])
                
                for col in range(1, 8):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = border_cell
                    cell.font = font_body
                    if col in [1, 7]:
                        cell.alignment = align_center_both
                    else:
                        cell.alignment = align_left_both
                        
                ws.row_dimensions[row_idx].height = 60
                
            ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=end_merge_row, end_column=1)
            ws.merge_cells(start_row=start_merge_row, start_column=2, end_row=end_merge_row, end_column=2)
            ws.merge_cells(start_row=start_merge_row, start_column=6, end_row=end_merge_row, end_column=6)
            ws.merge_cells(start_row=start_merge_row, start_column=7, end_row=end_merge_row, end_column=7)
            
            ws.cell(row=start_merge_row, column=1).font = font_bold_body
            ws.cell(row=start_merge_row, column=1).alignment = align_center_both
            ws.cell(row=start_merge_row, column=2).font = font_bold_body
            ws.cell(row=start_merge_row, column=6).alignment = align_left_both
            
            status_val = tc["status"]
            for r in range(start_merge_row, end_merge_row + 1):
                cell_status = ws.cell(row=r, column=7)
                if status_val == "Pass":
                    cell_status.fill = fill_pass
                    cell_status.font = font_status_pass
                else:
                    cell_status.fill = fill_fail
                    cell_status.font = font_status_fail
                cell_status.alignment = align_center_both
                cell_status.border = border_cell
                
            current_row += steps_count

        column_widths = {1: 15, 2: 25, 3: 35, 4: 55, 5: 25, 6: 30, 7: 15}
        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

    try:
        wb.save(EXCEL_FILENAME)
        print(f"[OK] Đã xuất báo cáo Excel thành công: '{os.path.abspath(EXCEL_FILENAME)}'")
        # Tự động mở file Excel sau khi tạo xong trên Windows
        try:
            os.startfile(EXCEL_FILENAME)
            print(f"[+] Tự động mở báo cáo Excel: '{EXCEL_FILENAME}'")
        except Exception as open_err:
            print(f"[!] Không thể tự động mở báo cáo Excel: {open_err}")
    except Exception as e:
        print(f"[ERROR] Không thể lưu file Excel: {e}")

if __name__ == "__main__":
    main()
